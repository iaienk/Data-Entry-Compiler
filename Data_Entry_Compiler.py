import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Font

# 📂 Percorsi dei file
# 📌 Ricava la cartella in cui si trova *questo* script
script_folder = os.path.dirname(__file__)

# 📌 Costruiamo i path relativi
sequence_file_path = os.path.join(script_folder, "Sequenza_Giornaliera.xlsx")
template_file_path = os.path.join(script_folder, "MoQ_10_Data entry.xlsx")

# 🔍 Controllo file
if not os.path.exists(sequence_file_path):
    print(f"❌ ERRORE: Il file {sequence_file_path} non esiste!")
    exit()

# 📖 Leggiamo Sequenza_Giornaliera.xlsx
sequence_df = pd.read_excel(sequence_file_path, sheet_name=0)

# 🔎 Colonne richieste
required_columns = [
    "Tipologia in salita", "Progressivo", "Data", "Linee", 
    "Tipologia in discesa", "Bus", "Direzione",
    "Fermata di salita", "Fermata di discesa", "Meteo"
]
# Aggiungiamo le nuove colonne "ID" e "Rilevatore" (colonna K e L)
required_columns.append("ID")
required_columns.append("Rilevatore")
# Aggiungiamo anche la nuova colonna "Tipo di giorno" (colonna M)
required_columns.append("Tipo di giorno")

missing_columns = [col for col in required_columns if col not in sequence_df.columns]
if missing_columns:
    print(f"❌ ERRORE: Manca una o più colonne: {missing_columns}")
    exit()

# 📅 Estrarre la data di riferimento dalla prima riga della colonna "Data"
date_reference = sequence_df["Data"].dropna().iloc[0]
try:
    date_reference = pd.to_datetime(date_reference, format="%d/%m/%Y", errors='coerce')
except:
    date_reference = datetime.today()

if pd.isna(date_reference):
    date_reference = datetime.today()

formatted_date = date_reference.strftime('%d/%m/%Y')
file_date_format = date_reference.strftime('%Y%m%d')
# 📂 Output in cartella script
# 📌 Estrai l'ID e determina il prefisso del file (01_ o 02_)
my_id = str(sequence_df["ID"].dropna().iloc[0])  # Assicuriamoci che sia stringa
file_prefix = "01_" if my_id.startswith("1") else "02_"

# 📌 Nome file con prefisso basato sull'ID
output_file_name = f"{file_prefix}{file_date_format}.xlsx"
output_file_path = os.path.join(script_folder, output_file_name)

print(f"📅 Data di riferimento: {formatted_date}")
print(f"📂 Il file verrà salvato come: {output_file_name}")

# 📌 Dizionario abbreviazioni → nomi reali dei fogli
abbreviazioni = {
    "E": "BASE CON ELETTRONICA",
    "PE": "BASE CON ELETTRONICA+PENSILINA",
    "NT": "BASE CON NUOVO TIPO",
    "NT+P": "BASE CON NUOVO TIPO+PENSILINA",
    "P": "BASE CON SOLO PENSILINA",
    "Provv": "BASE CON PROVVISORIA",
    "S": "BASE CON ELETTRONICA SOLARE",
}

# 📌 Estrarre linee monitorare
linee_monitorare = sequence_df["Linee"].dropna().values
if len(linee_monitorare) < 2:
    print("❌ ERRORE: Servono almeno due linee monitorare!")
    exit()

first_monitoring_line = linee_monitorare[0]
second_monitoring_line = linee_monitorare[1]

# 📌 Estrarre bus e direzioni
bus_numbers = sequence_df["Bus"].dropna().tolist()
directions = sequence_df["Direzione"].dropna().tolist()
if len(bus_numbers) < 10 or len(directions) < 10:
    print("❌ ERRORE: Servono almeno 10 valori di bus e direzioni!")
    exit()

# 📌 Estrarre fermata di salita, discesa e meteo
fermate_salita = sequence_df["Fermata di salita"].dropna().tolist()
fermate_discesa = sequence_df["Fermata di discesa"].dropna().tolist()
meteo = sequence_df["Meteo"].dropna().iloc[0]
if len(fermate_salita) < 10 or len(fermate_discesa) < 10:
    print("❌ ERRORE: Servono almeno 10 fermate di salita e di discesa!")
    exit()

# 📌 Estrarre ID e Rilevatore (colonne K e L)
my_id = sequence_df["ID"].dropna().iloc[0]  
rilevatore = sequence_df["Rilevatore"].dropna().iloc[0]

# 📌 Estrarre il "Tipo di giorno" (colonna M2)
tipo_giorno = sequence_df["Tipo di giorno"].dropna().iloc[0]

# 📌 Sequenze complete (10) e ridotte
sequence_list = sequence_df[["Tipologia in salita", "Progressivo"]].dropna(subset=["Tipologia in salita"]).values.tolist()
sequence_list = [(abbreviazioni.get(seq, seq), prog) for seq, prog in sequence_list]
sequence_ridotte = sequence_df["Tipologia in discesa"].dropna().tolist()
sequence_ridotte = [abbreviazioni.get(seq, seq) for seq in sequence_ridotte]

# 📌 righe_per_base
righe_per_base = {
    "BASE CON ELETTRONICA": (8, 20, [15,16]),
    "BASE CON ELETTRONICA+PENSILINA": (8, 27, [22,23]),
    "BASE CON NUOVO TIPO": (8, 22, [18,19]),
    "BASE CON NUOVO TIPO+PENSILINA": (8, 23, [19,20]),
    "BASE CON SOLO PENSILINA": (8, 23, [19,20]),
    "BASE CON PROVVISORIA": (8, 12, [10,11]),
    "BASE CON ELETTRONICA SOLARE": (8, 20, [15,16]),
}

# 📌 DIZIONARIO range+exclude fermate di salita
fermata_salita_ranges = {
    "BASE CON ELETTRONICA": {
        "range": (8, 20),
        "exclude": [15,16]
    },
    "BASE CON ELETTRONICA+PENSILINA": {
        "range": (8, 27),
        "exclude": [22,23]
    },
    "BASE CON NUOVO TIPO": {
        "range": (8, 22),
        "exclude": [18,19]
    },
    "BASE CON NUOVO TIPO+PENSILINA": {
        "range": (8, 23),
        "exclude": [19,20]
    },
    "BASE CON SOLO PENSILINA": {
        "range": (8, 23),
        "exclude": [19,20]
    },
    "BASE CON PROVVISORIA": {
        "range": (8, 12),
        "exclude": [10,11]
    },
    "BASE CON ELETTRONICA SOLARE": {
        "range": (8, 20),
        "exclude": [15,16]
    }
}

wb = load_workbook(template_file_path)
ws = wb["Data_Entry"]

def copia_bordi_traslati(source_cell, dest_ws, dest_row, dest_col):
    if source_cell.border and dest_row > 2:
        new_border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
            diagonal=source_cell.border.diagonal,
            diagonal_direction=source_cell.border.diagonal_direction,
            outline=source_cell.border.outline,
            vertical=source_cell.border.vertical,
            horizontal=source_cell.border.horizontal
        )
        cell_dest_up = dest_ws.cell(row=dest_row, column=dest_col)
        cell_dest_up.border = new_border

def forza_bold_se_sorgente_bold(source_cell, dest_cell):
    if source_cell.font and source_cell.font.bold:
        dest_cell.font = dest_cell.font.copy(bold=True)

def trova_prima_riga_vuota():
    row = 2
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        if all(cell.value is None for cell in r):
            return row
        row += 1
    return row

# -----------------------------------------------------------
#       SEQUENZE COMPLETE
# -----------------------------------------------------------
first_empty_row = trova_prima_riga_vuota()

for sheet_name, progressive_number in sequence_list:
    try:
        df = pd.read_excel(template_file_path, sheet_name=sheet_name, usecols=[0,1,2]).dropna(how="all")
    except Exception as e:
        print(f"⚠️ Errore nel leggere '{sheet_name}': {e}")
        continue

    monitoring_line = first_monitoring_line if progressive_number <= 5 else second_monitoring_line

    bus_index = int(progressive_number) - 1
    bus_number = bus_numbers[bus_index]
    direction  = directions[bus_index]
    fermata_salita_corrente = fermate_salita[bus_index]
    fermata_discesa_corrente = fermate_discesa[bus_index]

    total_rows = len(df)
    start_bus_direction = total_rows - 43

    try:
        ws_sorg = wb[sheet_name]
    except:
        ws_sorg = None

    # Carico range+exclude per fermata salita
    if sheet_name in fermata_salita_ranges:
        min_riga, max_riga = fermata_salita_ranges[sheet_name]["range"]
        exclude_list = fermata_salita_ranges[sheet_name]["exclude"]
    else:
        min_riga, max_riga = (1,9999)
        exclude_list = []

    for row_index, row_data in df.iterrows():
        # ID e Rilevatore
        ws[f"A{first_empty_row}"] = my_id
        ws[f"B{first_empty_row}"] = rilevatore

        ws[f"I{first_empty_row}"] = monitoring_line
        ws[f"N{first_empty_row}"] = int(progressive_number)
        ws[f"C{first_empty_row}"] = row_data.iloc[0]
        ws[f"P{first_empty_row}"] = row_data.iloc[1]
        ws[f"Q{first_empty_row}"] = row_data.iloc[2]
        ws[f"D{first_empty_row}"] = formatted_date

        # Meteo in colonna E
        ws[f"E{first_empty_row}"] = meteo

        # ❗️ Copia "Tipo di giorno" (colonna M in “Sequenza_Giornaliera”) → colonna F
        ws[f"F{first_empty_row}"] = tipo_giorno

        actual_row = row_index + 1

        # Bus e direzione nelle ultime 43 righe
        if row_index >= start_bus_direction:
            ws[f"K{first_empty_row}"] = bus_number
            ws[f"L{first_empty_row}"] = direction

        # Se è in exclude_list => bus/direzione lo stesso
        if actual_row in exclude_list:
            ws[f"K{first_empty_row}"] = bus_number
            ws[f"L{first_empty_row}"] = direction

        # Se rientra nel range => fermata salita in colonna J
        if min_riga <= actual_row <= max_riga and actual_row not in exclude_list:
            ws[f"J{first_empty_row}"] = fermata_salita_corrente

        # Se ultima riga => fermata discesa in col M
        if row_index == len(df) - 1:
            ws[f"M{first_empty_row}"] = fermata_discesa_corrente

        # Bordi + Bold
        if ws_sorg:
            sorg_row = row_index + 2
            col_map = {1:3, 2:16, 3:17}
            for src_col, dst_col in col_map.items():
                cell_da = ws_sorg.cell(row=sorg_row, column=src_col)
                copia_bordi_traslati(cell_da, ws, first_empty_row, dst_col)
                cell_a = ws.cell(row=first_empty_row, column=dst_col)
                forza_bold_se_sorgente_bold(cell_da, cell_a)

        first_empty_row += 1

# -----------------------------------------------------------
#       SEQUENZE RIDOTTE
# -----------------------------------------------------------
first_empty_row = trova_prima_riga_vuota()
progressive_number_ridotto = 1

for base_name in sequence_ridotte:
    if base_name not in righe_per_base:
        print(f"⚠️ ERRORE: '{base_name}' non esiste in righe_per_base!")
        continue

    # Scegliamo la "fermata_discesa" corrispondente al progressivo ridotto
    discesa_index = progressive_number_ridotto - 1
    fermata_discesa_ridotta = fermate_discesa[discesa_index]

    start_row, end_row, exclude_rows = righe_per_base[base_name]
    df = pd.read_excel(template_file_path, sheet_name=base_name, usecols=[0,1,2], skiprows=start_row - 1)
    df = df.iloc[:end_row - start_row + 1]

    exclude_indices = [r - start_row for r in exclude_rows if r>=start_row and r<=end_row]
    df = df.drop(exclude_indices, axis=0).dropna(how="all")

    try:
        ws_sorg_rid = wb[base_name]
    except:
        ws_sorg_rid = None

    for idx, row_data in enumerate(df.iterrows()):
        row_series = row_data[1]

        # ID e Rilevatore
        ws[f"A{first_empty_row}"] = my_id
        ws[f"B{first_empty_row}"] = rilevatore

        # Monitoring line (linea monitorare)
        ws[f"I{first_empty_row}"] = first_monitoring_line if progressive_number_ridotto <= 5 else second_monitoring_line

        # Progressivo ridotto
        ws[f"N{first_empty_row}"] = progressive_number_ridotto

        # D, C, P, Q
        ws[f"D{first_empty_row}"] = formatted_date
        ws[f"C{first_empty_row}"] = row_series.iloc[0]
        ws[f"P{first_empty_row}"] = row_series.iloc[1]
        ws[f"Q{first_empty_row}"] = row_series.iloc[2]

        # Meteo in colonna E
        ws[f"E{first_empty_row}"] = meteo

        # ❗️ Copia "Tipo di giorno" in colonna F
        ws[f"F{first_empty_row}"] = tipo_giorno

        # Inseriamo la fermata di discesa in colonna J per TUTTE le righe ridotte
        ws[f"J{first_empty_row}"] = fermata_discesa_ridotta

        # Bordi + bold
        if ws_sorg_rid:
            sorg_row = (start_row -1) + (idx+1) + 1
            col_map = {1:3, 2:16, 3:17}
            for src_col, dst_col in col_map.items():
                cell_da = ws_sorg_rid.cell(row=sorg_row, column=src_col)
                copia_bordi_traslati(cell_da, ws, first_empty_row, dst_col)
                cell_a = ws.cell(row=first_empty_row, column=dst_col)
                forza_bold_se_sorgente_bold(cell_da, cell_a)

        first_empty_row += 1

    progressive_number_ridotto += 1

def trova_ultima_riga_ridotte():
    last_row = 2
    ridotte_counter = 0
    ultima_riga_ridotte = 2
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=14, max_col=14):
        if r[0].value is not None:
            ridotte_counter += 1
            ultima_riga_ridotte = last_row
        if ridotte_counter == 10:
            while ws[f"N{ultima_riga_ridotte}"].value is not None:
                ultima_riga_ridotte += 1
            break
        last_row += 1
    return ultima_riga_ridotte - 1

def pulisci_righe_eccedenti():
    ultima_riga = trova_ultima_riga_ridotte()
    prima_riga_da_svuotare = ultima_riga + 1
    ultima_riga_da_svuotare = min(prima_riga_da_svuotare + 100, ws.max_row)
    print(f"🧹 Pulizia righe da {prima_riga_da_svuotare} a {ultima_riga_da_svuotare}...")
    for row in ws.iter_rows(min_row=prima_riga_da_svuotare, max_row=ultima_riga_da_svuotare, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None

pulisci_righe_eccedenti()

wb.save(output_file_path)
print(f"✅ File salvato con successo: {output_file_name}")
